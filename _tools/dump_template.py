# -*- coding: utf-8 -*-
"""检查 v5 报价模板结构 vs 真实 PI 布局。"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook

path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'quotation_template.xlsx')
wb = load_workbook(path, data_only=False)
print('sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('=' * 90)
    print('SHEET:', name, 'dims:', ws.dimensions, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
    print('merged:', sorted(str(m) for m in ws.merged_cells.ranges))
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                print('  R%dC%d: %s' % (r, c, str(v).replace('\n', ' | ')[:150]))
    # images
    try:
        imgs = getattr(ws, '_images', [])
        print('images:', len(imgs))
    except Exception:
        pass
wb.close()
