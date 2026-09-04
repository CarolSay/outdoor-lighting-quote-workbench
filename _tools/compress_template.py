# -*- coding: utf-8 -*-
"""压缩报价模板中的超大图片（降采样+重压缩），减小安装包体积。
- 最长边 > 1800px 的图缩到 1800px
- PNG 照片转 JPEG 需同步改扩展名 + 更新 drawing rels 引用 + [Content_Types].xml
- 仅当新文件明显更小才替换；备份原模板为 .bak
"""
import io
import os
import re
import shutil
import zipfile

from PIL import Image

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TPL = os.path.join(BASE, 'quotation_template.xlsx')
BAK = TPL + '.bak'
MAX_SIDE = 1800

src = zipfile.ZipFile(TPL)
contents = {i.filename: src.read(i.filename) for i in src.infolist()}
src.close()

rename_map = {}     # old media name -> new media name
new_contents = {}
for name, data in contents.items():
    if name.startswith('xl/media/') and name.lower().endswith(('.png', '.jpeg', '.jpg')):
        try:
            im = Image.open(io.BytesIO(data))
            w, h = im.size
            if max(w, h) > MAX_SIDE:
                scale = MAX_SIDE / max(w, h)
                im = im.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
            if im.mode in ('RGBA', 'P', 'LA', 'CMYK'):
                im = im.convert('RGB')
            buf = io.BytesIO()
            im.save(buf, 'JPEG', quality=82, optimize=True)
            nd = buf.getvalue()
            if len(nd) < len(data) * 0.7:            # 至少缩小 30% 才值得替换
                newname = re.sub(r'\.(png|jpe?g)$', '.jpeg', name, flags=re.I)
                if newname != name:
                    rename_map[name] = newname
                new_contents[newname] = nd
                print('替换 %s(%0.1fMB) -> %s(%0.2fMB)' % (name, len(data) / 1048576, newname, len(nd) / 1048576))
                continue
            elif max(w, h) > MAX_SIDE:
                # 尺寸超限但 JPEG 没小多少，仍存降采样 PNG
                buf2 = io.BytesIO()
                im.save(buf2, 'PNG', optimize=True)
                if len(buf2.getvalue()) < len(data) * 0.7:
                    new_contents[name] = buf2.getvalue()
                    print('替换 %s -> 降采样PNG %0.2fMB' % (name, len(buf2.getvalue()) / 1048576))
                    continue
        except Exception as e:
            print('跳过', name, e)
    new_contents[name] = data

# 更新引用了旧媒体名的 rels
for name in list(new_contents.keys()):
    if name.endswith('.rels'):
        txt = new_contents[name].decode('utf-8')
        orig = txt
        for old, new in rename_map.items():
            txt = txt.replace(os.path.basename(old), os.path.basename(new))
        if txt != orig:
            print('更新引用:', name)
            new_contents[name] = txt.encode('utf-8')

# [Content_Types].xml 确保 jpeg 默认类型
ct = new_contents['[Content_Types].xml'].decode('utf-8')
if 'Extension="jpeg"' not in ct:
    ct = ct.replace('<Types ', '<Types ', 1).replace(
        '</Types>', '<Default Extension="jpeg" ContentType="image/jpeg"/></Types>')
    print('补充 [Content_Types].xml jpeg 声明')
    new_contents['[Content_Types].xml'] = ct.encode('utf-8')

out = io.BytesIO()
zout = zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED, compresslevel=6)
for name, data in new_contents.items():
    zout.writestr(name, data)
zout.close()

old_size = os.path.getsize(TPL)
shutil.copy(TPL, BAK)
with open(TPL, 'wb') as f:
    f.write(out.getvalue())
new_size = os.path.getsize(TPL)
print('模板大小: %.2f MB -> %.2f MB' % (old_size / 1048576, new_size / 1048576))

# 验证 openpyxl 能正常打开（图片引用完整）
from openpyxl import load_workbook
wb = load_workbook(TPL)
imgs = wb.worksheets[0]._images
print('验证打开 OK, sheets:', wb.sheetnames, '图片数:', len(imgs))
