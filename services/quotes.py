# -*- coding: utf-8 -*-
"""报价与历史Excel导入/导出(从 V3 单文件拆出)。
含: Excel 解析(xlsx/xls)、去重、导入确认、报价单模板生成、原文件下载。
"""
import base64
import hashlib
import json
import os
import re
import shutil
import sqlite3
import subprocess
import tempfile
from datetime import datetime, timedelta

import config as C
from db import tx, query_all, query_one

# ---------- 基础文本工具 ----------
def cell_text(v):
    return '' if v is None else str(v).strip()


def _norm(v):
    return re.sub(r'\s+', '', cell_text(v)).lower()


def norm_date(v):
    if v is None or v == '':
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    txt = cell_text(v)
    m = re.match(r'^(\d{4})[\./-](\d{1,2})[\./-](\d{1,2})', txt)
    if m:
        return '%04d-%02d-%02d' % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return txt


def display_date(v):
    if not v:
        return ''
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', str(v))
    return '%d/%d/%d' % (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else str(v)


def to_num(v):
    if v is None or v == '':
        return 0
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0


# ---------- Excel/PDF 解析 ----------
_META_EMPTY = {'project_name': '', 'customer_text': '', 'quote_no': '', 'quote_date': '', 'provider': '',
               'contact': '', 'address': '', 'phone': ''}


def _classify_label(n):
    """表头标签归位（n 已 _norm：小写无空白，中文保留）。真实标签带中文注释，如 consignee（公司）。"""
    if 'date' in n or '日期' in n:
        return 'quote_date'
    if ('s/c' in n or '合同' in n) and ('no' in n or '编号' in n):
        return 'quote_no'
    if n.startswith('consignee') or n.startswith('customer') or n.startswith('client') or '客户' in n:
        return 'customer_text'
    if n.startswith('contact') or n.startswith('收货人') or '联系人' in n:
        return 'contact'
    if n.startswith('address') or '地址' in n:
        return 'address'
    if n.startswith('phone') or n.startswith('tel') or '电话' in n:
        return 'phone'
    if n.startswith('project') or '项目' in n:
        return 'project_name'
    if n.startswith('provider') or '供应商' in n:
        return 'provider'
    if n in ('pi',) or n.startswith('quot') or n.startswith('pi') or '报价' in n:
        return 'quote_no'
    return None


def _extract_grid(vals, title):
    """从单元格二维数组提取 (rows, meta)。xlsx/xls 共用。"""
    meta = dict(_META_EMPTY)
    rows = []
    for row in vals:
        ns = [_norm(v) for v in row]
        for i, n in enumerate(ns):
            if not n or i + 1 >= len(row):
                continue
            v = cell_text(row[i + 1])
            if not v:
                continue
            k = _classify_label(n)
            if k and not meta.get(k):
                meta[k] = norm_date(v) if k == 'quote_date' else v
    header, hr = None, -1
    for r, row in enumerate(vals):
        ns = [_norm(v) for v in row]
        aliases = {}
        for i, n in enumerate(ns):
            if n.startswith('itemname'):
                aliases.setdefault('itemname', i)
            elif n.startswith('description'):
                aliases.setdefault('description', i)
            elif n.startswith('quantity'):
                aliases.setdefault('quantity', i)
            elif n.startswith('unitprice'):
                aliases.setdefault('unitprice', i)
            elif n.startswith('amount'):
                aliases.setdefault('amount', i)
        if {'itemname', 'description', 'quantity', 'unitprice', 'amount'}.issubset(aliases):
            header, hr = aliases, r
            break
    if header is None:
        return rows, meta
    for r in range(hr + 1, len(vals)):
        row = vals[r]
        gv = lambda k: row[header[k]] if header.get(k) is not None and header.get(k) < len(row) else ''
        item = cell_text(gv('itemname'))
        desc = cell_text(gv('description'))
        qty = to_num(gv('quantity'))
        up = to_num(gv('unitprice'))
        amt = to_num(gv('amount'))
        if amt == 0 and qty and up:
            amt = qty * up
        if not any([item, desc, qty, up, amt]):
            continue
        if _norm(item) in ('total', 'subtotal', 'notes', 'remark', 'remarks', 'sum', '总额', 'commision',
                           'commission'):
            continue
        unit = 'pcs'
        qi, upi = header.get('quantity', -1), header.get('unitprice', -1)
        if qi >= 0 and qi + 1 < len(row) and qi + 1 != upi and cell_text(row[qi + 1]) \
                and not re.match(r'^-?\d+(?:\.\d+)?$', cell_text(row[qi + 1])):
            unit = cell_text(row[qi + 1])
        rows.append({'sheet_name': title, 'project_name': meta['project_name'], 'item_name': item,
                     'description': desc, 'quantity': qty, 'unit': unit, 'unit_price': up, 'amount': amt,
                     'source_row': r + 1, 'raw_json': json.dumps(row, ensure_ascii=False, default=str)})
    return rows, meta


def parse_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False, read_only=False)
    rows, meta = [], dict(_META_EMPTY)
    for ws in wb.worksheets:
        maxc = min(ws.max_column, 40)
        maxr = min(ws.max_row, 500)
        vals = [[ws.cell(r, c).value for c in range(1, maxc + 1)] for r in range(1, maxr + 1)]
        r2, m2 = _extract_grid(vals, ws.title)
        if r2:
            rows.extend(r2)
        for k, v in m2.items():
            if v and not meta.get(k):
                meta[k] = v
    return rows, meta


def parse_xls(path):
    """旧版 .xls：xlrd 读取（xls 专用），转成二维数组后复用 _extract_grid。"""
    import xlrd
    wb = xlrd.open_workbook(path)
    rows, meta = [], dict(_META_EMPTY)
    for ws in wb.sheets():
        if ws.nrows == 0:
            continue
        vals = []
        for r in range(min(ws.nrows, 500)):
            vals.append([ws.cell_value(r, c) if c < ws.ncols else '' for c in range(min(ws.ncols, 40))])
        r2, m2 = _extract_grid(vals, ws.name)
        if r2:
            rows.extend(r2)
        for k, v in m2.items():
            if v and not meta.get(k):
                meta[k] = v
    return rows, meta


def parse_pdf(path):
    """PDF：文本提取 → 分类（报价单/PI/合同）→ 尽力解析明细行。
    返回 dict: kind('pi'|'quotation'|'contract'|None), rows, meta, text。"""
    from pypdf import PdfReader
    rd = PdfReader(path)
    pages = []
    for p in rd.pages[:5]:
        try:
            pages.append(p.extract_text() or '')
        except Exception:
            pages.append('')
    text = '\n'.join(pages)
    meta = dict(_META_EMPTY)
    m = re.search(r'(?:S/?C\s*No\.?|合同编号)\s*[.：:（(]*\s*([A-Z0-9][A-Z0-9\-/]{4,40})', text, re.I)
    if m:
        meta['quote_no'] = m.group(1).strip()
    m = re.search(r'(?:S/?C\s*Date|日期)\s*[.：:（(]*\s*(\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2})', text, re.I)
    if m:
        meta['quote_date'] = norm_date(m.group(1))
    # PDF 格式：标签块(Consignee/Contact/Address/Phone)在前，值块在 Bank Info 之后
    # 值块格式：公司名 → 供应商信息(多行) → 联系人 → 地址 → 电话 → Sum
    m = re.search(r'SWIFT\s*CODE[^\n]*\n', text, re.I)
    if m:
        after_swift = text[m.end():]
        lines_after = [ln.strip() for ln in after_swift.splitlines() if ln.strip()]
        # 过滤掉供应商行(含SHENZHEN/Add:/Email:/T：/SWIFT等)
        cust_lines = [ln for ln in lines_after
                      if not re.match(r'^(?:SHENZHEN|Add:|Email:|M:|T[：:π]|Benificiary|Bank\s|SWIFT)', ln, re.I)
                      and not re.match(r'^(?:Sum|总|Project|PO\b|PI\b)', ln, re.I)
                      and not ln.startswith('（')
                      and not re.match(r'^[（(]', ln)
                      and len(ln) >= 3]
        if len(cust_lines) >= 1:
            meta['customer_text'] = cust_lines[0]
        if len(cust_lines) >= 2:
            meta['contact'] = cust_lines[1]
        if len(cust_lines) >= 3:
            meta['address'] = cust_lines[2]
        if len(cust_lines) >= 4:
            phone_m = re.search(r'(\d[\d\s\-+()]{6,20})', cust_lines[3])
            if phone_m:
                meta['phone'] = phone_m.group(1).strip()
    # 从 PI/项目 行提取项目名
    m = re.search(r'(?:Project|项目名称)\s*[（(]([^)）]+)[)）]\s*\n([^\n]{3,80})', text, re.I)
    if m:
        meta['project_name'] = m.group(2).strip()
    has_bank = bool(re.search(r'benificiary|bank\s+account|swift', text, re.I))
    is_contract = bool(re.search(r'采购合同|销售合同|购销合同|contract|agreement', text, re.I)) and not has_bank \
        and not re.search(r'proforma|quotation|报价单|pi\b', text, re.I)
    if is_contract:
        return {'kind': 'contract', 'rows': [], 'meta': meta, 'text': text}
    kind = None
    if re.search(r'proforma\s*invoice|\bPI\b|pi\s*for', text, re.I) or has_bank:
        kind = 'pi'
    elif re.search(r'quotation|报价单|quote', text, re.I):
        kind = 'quotation'
    # 明细行解析：支持单行和多行两种格式
    rows = []
    lines = text.splitlines()
    # 单行格式：序号 名称(多空格)描述 数量 单位 $单价 $金额
    line_re = re.compile(r'^\s*(\d{1,3})\s+(\S[^\n]*?)\s{2,}(.+?)\s+(\d+(?:,\d{3})*(?:\.\d+)?)\s*'
                         r'(pcs|set|m|pc|pcs\.?)?\s+\$?([\d,.]+)\s+\$?([\d,]+\.\d+)\s*$', re.I)
    # 多行格式：序号 + 名称 单独一行，后续行是描述，最后行是 数量 单位 $单价 $金额
    qty_line_re = re.compile(r'^\s*(\d+(?:,\d{3})*(?:\.\d+)?)\s*(pcs|set|m|pc|pcs\.?)?\s+\$?([\d,.]+)\s+\$?([\d,]+\.\d+)\s*$', re.I)
    item_start_re = re.compile(r'^\s*(\d{1,3})\s+(\S[^\n]*?)\s*$', re.I)
    i = 0
    while i < len(lines):
        ln = lines[i].strip()
        # 尝试单行匹配
        mm = line_re.match(ln)
        if mm:
            no, name, desc, qty, unit, up, amt = mm.groups()
            rows.append({'sheet_name': 'PDF', 'project_name': meta.get('project_name', ''),
                         'item_name': name.strip(), 'description': desc.strip(), 'quantity': to_num(qty),
                         'unit': unit or 'pcs', 'unit_price': to_num(up), 'amount': to_num(amt.replace(',', '')),
                         'source_row': i, 'raw_json': json.dumps({'line': ln}, ensure_ascii=False)})
            i += 1
            continue
        # 尝试多行匹配：序号+名称行
        ms = item_start_re.match(ln)
        if ms and i + 1 < len(lines):
            no, name = ms.groups()
            # 收集描述行直到找到数量行
            desc_parts = []
            j = i + 1
            found_qty = None
            while j < len(lines) and j < i + 10:
                ml = lines[j].strip()
                mq = qty_line_re.match(ml)
                if mq:
                    qty, unit, up, amt = mq.groups()
                    found_qty = (qty, unit or 'pcs', up, amt)
                    break
                if ml and not re.match(r'^(?:bank|benificiary|swift|总|sum)', ml, re.I):
                    desc_parts.append(ml)
                else:
                    break
                j += 1
            if found_qty:
                qty, unit, up, amt = found_qty
                rows.append({'sheet_name': 'PDF', 'project_name': meta.get('project_name', ''),
                             'item_name': name.strip(), 'description': ' '.join(desc_parts).strip(),
                             'quantity': to_num(qty), 'unit': unit, 'unit_price': to_num(up),
                             'amount': to_num(amt.replace(',', '')),
                             'source_row': i, 'raw_json': json.dumps({'lines': lines[i:j+1]}, ensure_ascii=False)})
                i = j + 1
                continue
        i += 1
    if rows and not kind:
        kind = 'pi' if has_bank else 'quotation'
    return {'kind': kind, 'rows': rows, 'meta': meta, 'text': text[:3000]}


def classify_doc(name, rows, meta):
    """文件级分类：文件名优先，其次内容特征（PI 描述含 hs code）。返回 'pi' | 'quotation'。"""
    nl = (name or '').lower()
    base = os.path.splitext(os.path.basename(nl))[0]
    if re.match(r'^pi[\s_-]|pi\s*for|proforma', base):
        return 'pi'
    if 'quotation' in base or '报价' in base or 'quote' in base:
        return 'quotation'
    for x in rows:
        if 'hs code' in (x.get('description') or '').lower():
            return 'pi'
    if re.search(r'\bpi\b', nl):
        return 'pi'
    return 'quotation'


def parse_blob(name, blob):
    """解析文件 → (rows, meta, doc_type)。
    doc_type: 'pi' / 'quotation' / 'contract' / None(需按名称与内容分类)。contract 时 rows=[]。"""
    ext = os.path.splitext(name)[1].lower()
    fd, temp = tempfile.mkstemp(suffix=ext)
    os.close(fd)
    try:
        with open(temp, 'wb') as f:
            f.write(blob)
        if ext == '.pdf':
            pr = parse_pdf(temp)
            if pr['kind'] == 'contract':
                return [], pr['meta'], 'contract'
            return pr['rows'], pr['meta'], pr['kind']
        rows, meta = parse_xlsx(temp) if ext == '.xlsx' else parse_xls(temp)
        return rows, meta, None
    finally:
        try:
            os.remove(temp)
        except Exception:
            pass


# ---------- 客户/供应商 获取或创建(导入用) ----------
def get_or_create_provider(c, text):
    txt = (text or '').strip()
    if not txt:
        return c.execute('SELECT id FROM providers WHERE active=1 ORDER BY id LIMIT 1').fetchone()['id']
    first = (txt.splitlines()[0].strip() or txt[:120]).strip()
    r = c.execute('SELECT id FROM providers WHERE lower(provider_name)=lower(?) LIMIT 1', (first,)).fetchone()
    if r:
        return r['id']
    code = 'P' + hashlib.sha1(txt.encode('utf-8', 'ignore')).hexdigest()[:8].upper()
    c.execute('INSERT OR IGNORE INTO providers(provider_code,provider_name,provider_info,active) VALUES (?,?,?,1)',
              (code, first, txt))
    return c.execute('SELECT id FROM providers WHERE provider_code=?', (code,)).fetchone()['id']


def get_or_create_customer(c, text, contact='', address='', phone=''):
    """按公司名获取或创建客户；已存在时补齐空的联系字段。"""
    name = (text or 'Imported Customer').strip()
    parts = [x.strip() for x in re.split(r'[\r\n]+', name) if x.strip()]
    company = parts[0] if parts else name
    r = c.execute('SELECT id FROM customers WHERE lower(company)=lower(?) LIMIT 1', (company,)).fetchone()
    if r:
        cid = r['id']
        sets, vals = [], []
        for col, v in (('contact', contact), ('address', address), ('phone', phone)):
            if v and not c.execute('SELECT 1 FROM customers WHERE id=? AND (%s IS NOT NULL AND %s<>"")'
                                   % (col, col), (cid,)).fetchone():
                sets.append('%s=?' % col)
                vals.append(v)
        if sets:
            c.execute('UPDATE customers SET %s,updated_at=CURRENT_TIMESTAMP WHERE id=?'
                      % ','.join(sets), vals + [cid])
        return cid
    code = 'CIMP' + hashlib.sha1(company.encode('utf-8', 'ignore')).hexdigest()[:8].upper()
    c.execute('INSERT OR IGNORE INTO customers(customer_code,company,country,contact,address,phone,currency,active) '
              'VALUES (?,?,?,?,?,?,?,1)',
              (code, company, '', contact or '; '.join(parts[1:]) or None, address or None, phone or None, 'USD'))
    return c.execute('SELECT id FROM customers WHERE customer_code=?', (code,)).fetchone()['id']


# ---------- 导入时自动入库产品 ----------
def _norm_prod_text(v):
    t = re.sub(r'\s+', ' ', cell_text(v)).strip()
    return t or None


def _gen_product_code(c):
    """生成唯一产品编码(与 routes.core._gen_code 同风格, 编码自校验唯一)。"""
    base = datetime.now().strftime('%Y%m%d%H%M%S%f')[-10:]
    code = 'CM-P-' + base
    if not c.execute('SELECT 1 FROM products WHERE product_code=?', (code,)).fetchone():
        return code
    n = c.execute('SELECT COUNT(*) FROM products').fetchone()[0]
    return 'CM-P-%s%03d' % (datetime.now().strftime('%H%M%S'), n + 1)


def _is_blank_line(qty, price, amount):
    """无数量/单价/金额的明细行(如页脚文字被误识别)，不入库为产品。"""
    return to_num(qty) <= 0 and to_num(price) <= 0 and to_num(amount) <= 0


def _guess_category(name, desc):
    """从名称/描述猜测产品类别。"""
    t = ((name or '') + ' ' + (desc or '')).lower()
    rules = [('线条灯', 'linear'), ('洗墙灯', 'washer'), ('wall washer', 'washer'),
             ('投光灯', 'flood'), ('点光源', 'pixel'), ('point light', 'pixel'), ('dot', 'pixel'),
             ('控制器', 'controller'), ('sub controller', 'controller'), ('main controller', 'controller'),
             ('touch panel', 'controller'), ('电源', 'power supply'), ('driver', 'power supply'),
             ('power supply', 'power supply'), ('线缆', 'cable'), ('cable', 'cable'),
             ('连接器', 'accessory'), ('connector', 'accessory'), ('accessorry', 'accessory'),
             ('accessory', 'accessory'), ('stopper', 'accessory')]
    for cn, en in rules:
        if cn in t:
            return {'linear': '线条灯', 'washer': '洗墙灯', 'flood': '投光灯', 'pixel': '点光源',
                    'controller': '控制器', 'power supply': '电源', 'cable': '线缆',
                    'accessory': '配件'}[en]
    return '其他'


# 描述解析出的字段 → products 固定列映射
_PARSE_COL_MAP = ['model', 'voltage', 'power', 'cct_color', 'cct', 'control', 'ip_rating', 'beam_angle',
                  'length_size', 'led_count', 'pixel_count', 'led_chip', 'material', 'body_color',
                  'hs_code', 'notes', 'ext1', 'ext2', 'ext3']


def get_or_create_product(c, item_name, description, unit_price, source='', doc_type='quotation'):
    """把报价明细当作产品入库：
    1) 描述解析出型号 → 按 model 匹配，命中复用（补空字段+更新价格）
    2) 按 商品名称+商品描述 匹配
    3) 新建：解析字段写入固定列，其余进 spec_json，描述按 PI 风格重建
    返回 (product_id, is_new)。"""
    from services.spec_fields import SPEC_KEYS, parse_description_safe, build_description
    nm = _norm_prod_text(item_name)
    desc = _norm_prod_text(description)
    if not nm:
        nm = (desc[:120].strip()) if desc else None   # 无名称时以描述开头兜底
    if not nm and not desc:
        return None, False
    fields = parse_description_safe(desc)
    price = to_num(unit_price)
    # 1) 型号优先匹配
    model = fields.get('model')
    if model:
        ex = c.execute('SELECT * FROM products WHERE active=1 AND model=? ORDER BY id LIMIT 1',
                       (model,)).fetchone()
        if ex:
            sets, vals = [], []
            for col in _PARSE_COL_MAP:
                v = fields.get(col)
                if v and not ex[col]:
                    sets.append('%s=?' % col)
                    vals.append(v)
            if price > 0 and not ex['standard_price_usd']:
                sets.append('standard_price_usd=?')
                vals.append(price)
            if sets:
                c.execute('UPDATE products SET %s,updated_at=CURRENT_TIMESTAMP WHERE id=?'
                          % ','.join(sets), vals + [ex['id']])
            return ex['id'], False
    # 2) 名称+描述匹配
    ex = c.execute('SELECT id FROM products WHERE active=1 AND product_name IS ? AND description IS ? LIMIT 1',
                   (nm, desc)).fetchone()
    if ex:
        return ex['id'], False
    # 3) 新建
    code = _gen_product_code(c)
    spec = {k: fields[k] for k in SPEC_KEYS if fields.get(k)}
    core_vals = {k: fields.get(k) for k in _PARSE_COL_MAP}
    notes = core_vals.pop('notes', None) or None
    new_desc = build_description(fields) or desc
    c.execute('INSERT INTO products(product_code,series,model,product_name,description,'
              'voltage,power,cct_color,cct,control,ip_rating,beam_angle,length_size,led_count,pixel_count,'
              'led_chip,material,body_color,hs_code,currency,category,standard_price_usd,notes,spec_json,'
              'ext1,ext2,ext3,active) '
              'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)',
              (code, 'Imported', core_vals.get('model'), nm, new_desc,
               core_vals.get('voltage'), core_vals.get('power'), core_vals.get('cct_color'), core_vals.get('cct'),
               core_vals.get('control'), core_vals.get('ip_rating'), core_vals.get('beam_angle'),
               core_vals.get('length_size'), core_vals.get('led_count'), core_vals.get('pixel_count'),
               core_vals.get('led_chip'), core_vals.get('material'), core_vals.get('body_color'),
               core_vals.get('hs_code'), 'USD', _guess_category(nm, desc),
               price if price > 0 else 0, notes,
               json.dumps(spec, ensure_ascii=False) if spec else None,
               core_vals.get('ext1'), core_vals.get('ext2'), core_vals.get('ext3')))
    pid = c.execute('SELECT id FROM products WHERE product_code=?', (code,)).fetchone()['id']
    return pid, True


def backfill_products_from_items():
    """一次性回填：把历史 quotation_items 中未关联(或关联失效)的明细按 名称+描述 入库并关联 product_id。"""
    items = query_all('SELECT id,product_id,product_name,description,qty,unit_price_usd,amount_usd '
                      'FROM quotation_items ORDER BY id')
    stats = {'created': 0, 'linked': 0, 'skipped_blank': 0}

    def work(c):
        for it in items:
            if _is_blank_line(it.get('qty'), it.get('unit_price_usd'), it.get('amount_usd')):
                # 页脚/空行：解绑并跳过，不作为产品入库
                if it['product_id']:
                    c.execute('UPDATE quotation_items SET product_id=NULL WHERE id=?', (it['id'],))
                stats['skipped_blank'] += 1
                continue
            if it['product_id'] and c.execute('SELECT 1 FROM products WHERE id=?', (it['product_id'],)).fetchone():
                continue
            pid, is_new = get_or_create_product(c, it['product_name'], it['description'],
                                                it.get('unit_price_usd') or 0, '历史报价回填')
            if pid and pid != it['product_id']:
                c.execute('UPDATE quotation_items SET product_id=? WHERE id=?', (pid, it['id']))
                stats['linked'] += 1
                if is_new:
                    stats['created'] += 1
        # 清理导入产生但已不被任何明细引用的产品(页脚垃圾行)
        c.execute("DELETE FROM products WHERE series='Imported' AND id NOT IN "
                  "(SELECT product_id FROM quotation_items WHERE product_id IS NOT NULL)")
        return stats
    return tx(work)


# ---------- 导入核心 ----------
def import_blob(c, name, relpath, blob, rows, meta, sha, doc_type='quotation'):
    old = c.execute('SELECT id FROM import_files WHERE sha256=?', (sha,)).fetchone()
    if old:
        return {'status': 'duplicate', 'file_id': old['id'], 'message': '文件内容已存在，跳过重复导入'}
    fid = c.execute(
        'INSERT INTO import_files(file_name,relative_path,file_ext,sha256,file_size,status,workbook_blob,doc_type,'
        'quote_no,quote_date,customer_text,project_name) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
        (name, relpath, os.path.splitext(name)[1].lower(), sha, len(blob), '已导入', sqlite3.Binary(blob),
         doc_type, meta.get('quote_no', ''), meta.get('quote_date', ''), meta.get('customer_text', ''),
         meta.get('project_name', ''))).lastrowid
    proj = meta.get('project_name', '')
    cust = meta.get('customer_text', '')
    qno = meta.get('quote_no', '') or 'IMP-' + sha[:12].upper()
    qdate = meta.get('quote_date', '') or datetime.now().strftime('%Y-%m-%d')
    customer_id = get_or_create_customer(c, cust, meta.get('contact', ''), meta.get('address', ''),
                                         meta.get('phone', ''))
    provider_id = get_or_create_provider(c, meta.get('provider', ''))
    project_id = None
    if proj:
        pr = c.execute('SELECT id FROM projects WHERE customer_id=? AND project_name=? LIMIT 1',
                       (customer_id, proj)).fetchone()
        if pr:
            project_id = pr['id']
        else:
            pcode = 'PRJIMP-' + sha[:10].upper()
            c.execute('INSERT OR IGNORE INTO projects(project_code,customer_id,project_name,project_type,status,'
                      'quotation_no,modified_at) VALUES (?,?,?,?,?,?,CURRENT_TIMESTAMP)',
                      (pcode, customer_id, proj, 'Imported Quotation', '报价中', qno))
            project_id = c.execute('SELECT id FROM projects WHERE project_code=?', (pcode,)).fetchone()['id']
    qrow = c.execute('SELECT id FROM quotations WHERE quote_no=?', (qno,)).fetchone()
    total = sum(x['amount'] for x in rows)
    expiry = ''
    if re.match(r'^\d{4}-\d{2}-\d{2}$', qdate):
        expiry = (datetime.strptime(qdate, '%Y-%m-%d') + timedelta(days=30)).strftime('%Y-%m-%d')
    if qrow:
        qid = qrow['id']
        c.execute('UPDATE quotations SET quote_date=?,customer_id=?,project_id=?,provider_id=?,expiry_date=?,'
                  'total_usd=?,updated_at=CURRENT_TIMESTAMP WHERE id=?',
                  (qdate, customer_id, project_id, provider_id, expiry, total, qid))
        c.execute('DELETE FROM quotation_items WHERE quotation_id=?', (qid,))
    else:
        qid = c.execute('INSERT INTO quotations(quote_no,quote_date,customer_id,project_id,provider_id,expiry_date,'
                        'currency,status,total_usd,notes,is_formal) VALUES (?,?,?,?,?,?,?,?,?,?,1)',
                        (qno, qdate, customer_id, project_id, provider_id, expiry, 'USD', '正式版本', total,
                         '导入来源：' + relpath)).lastrowid
    new_products = 0
    for n, x in enumerate(rows, 1):
        pid = None
        if not _is_blank_line(x['quantity'], x['unit_price'], x['amount']):
            pid, is_new = get_or_create_product(c, x['item_name'], x['description'], x['unit_price'],
                                                relpath, doc_type)
            if is_new:
                new_products += 1
        c.execute('INSERT INTO quotation_items(quotation_id,item_no,product_id,product_name,description,qty,unit,'
                  'unit_price_usd,amount_usd) VALUES (?,?,?,?,?,?,?,?,?)',
                  (qid, n, pid, x['item_name'], x['description'], x['quantity'], x.get('unit', 'pcs'),
                   x['unit_price'], x['amount']))
        c.execute('INSERT INTO imported_quote_rows(import_file_id,sheet_name,project_name,item_name,description,'
                  'quantity,unit,unit_price,amount,source_row,raw_json) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                  (fid, x['sheet_name'], proj, x['item_name'], x['description'], x['quantity'], x.get('unit', 'pcs'),
                   x['unit_price'], x['amount'], x['source_row'], x['raw_json']))
    c.execute('UPDATE quotations SET provider_id=COALESCE(provider_id,?), expiry_date=COALESCE(expiry_date,?) '
              'WHERE id=?', (provider_id, expiry, qid))
    # 报价历史：记录来源类型与文件位置（只查询展示用）
    # 相同报价号但不同文件（不同来源）时仍创建历史记录
    source_type = '导入-PI' if doc_type == 'pi' else '导入-报价单'
    if not c.execute('SELECT 1 FROM quotation_history WHERE quotation_id=? AND source_file=? LIMIT 1',
                     (qid, relpath)).fetchone():
        c.execute('INSERT INTO quotation_history(quotation_id,quote_no,quote_date,customer_id,project_id,'
                  'provider_id,expiry_date,currency,total_usd,status,notes,source_type,source_file) '
                  'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
                  (qid, qno, qdate, customer_id, project_id, provider_id, expiry, 'USD', total, '正式版本',
                   '导入来源：' + relpath, source_type, relpath))
        c.execute('UPDATE quotation_history SET total_usd=?,quote_date=?,notes=? WHERE quotation_id=? AND source_file=?',
                  (total, qdate, '导入来源：' + relpath, qid, relpath))
    return {'status': 'imported', 'file_id': fid, 'quotation_id': qid, 'quote_no': qno,
            'rows': len(rows), 'products_new': new_products, 'doc_type': doc_type}


def scan_files(scanid, files):
    """files: [{file_name, relative_path, content_b64}] -> 逐文件解析并落 import_scan_files"""
    results, seen = [], set()
    for item in files:
        name = item.get('file_name', '')
        rel = item.get('relative_path', '')
        ext = os.path.splitext(name)[1].lower()
        try:
            blob = base64.b64decode(item.get('content_b64', ''))
        except Exception:
            blob = b''
        sha = hashlib.sha256(blob).hexdigest()
        status, msg, rows, meta, doc_type = '新文件', '等待确认导入', [], {}, ''
        try:
            if sha in seen or query_one('SELECT id FROM import_files WHERE sha256=?', (sha,)):
                status = '重复'
                msg = '数据库中已有相同文件内容' if sha not in seen else '本次扫描中重复'
            else:
                seen.add(sha)
                rows, meta, doc_type = parse_blob(name, blob)
                if not doc_type:
                    doc_type = classify_doc(name, rows, meta)
                if doc_type == 'contract':
                    status = '跳过'
                    msg = '识别为合同文件，不导入'
                elif not rows:
                    status = '解析失败'
                    msg = '未找到 ItemName / Description / Quantity / UnitPrice / Amount 明细表'
        except Exception as e:
            status = '解析失败'
            msg = str(e)

        def ins(c):
            return c.execute(
                'INSERT INTO import_scan_files(scan_id,file_name,relative_path,file_ext,file_size,sha256,status,'
                'message,workbook_blob,doc_type,quote_no,quote_date,customer_text,customer_contact,customer_address,'
                'customer_phone,project_name,row_count) '
                'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                (scanid, name, rel, ext, len(blob), sha, status, msg,
                 sqlite3.Binary(blob) if status == '新文件' else None,
                 doc_type or '', meta.get('quote_no', ''), meta.get('quote_date', ''),
                 meta.get('customer_text', ''), meta.get('contact', ''), meta.get('address', ''),
                 meta.get('phone', ''), meta.get('project_name', ''), len(rows))).lastrowid
        sid = tx(ins)
        results.append({'id': sid, 'file_name': name, 'relative_path': rel, 'status': status,
                        'message': msg, 'row_count': len(rows), 'doc_type': doc_type or ''})
    return results


def confirm_import(ids):
    done, errors = 0, []
    for sid in ids:
        sf = query_one('SELECT * FROM import_scan_files WHERE id=?', (sid,))
        if not sf or sf['status'] != '新文件':
            continue
        try:
            rows, meta, doc_type = parse_blob(sf['file_name'], sf['workbook_blob'])
            if doc_type == 'contract':
                tx(lambda c: c.execute("UPDATE import_scan_files SET status='跳过',message=? WHERE id=?",
                                       ('识别为合同文件，不导入', sid)))
                continue

            def f(c):
                dt = doc_type or sf['doc_type'] or classify_doc(sf['file_name'], rows, meta) or 'quotation'
                return import_blob(c, sf['file_name'], sf['relative_path'], sf['workbook_blob'],
                                   rows, meta, sf['sha256'], doc_type=dt)
            r = tx(f)
            if r.get('status') == 'imported':
                done += 1
                tx(lambda c: c.execute("UPDATE import_scan_files SET status='已导入',message=? WHERE id=?",
                                       ('已写入报价列表；%d条明细，新增产品%d个' % (len(rows), r.get('products_new', 0)), sid)))
            else:
                tx(lambda c: c.execute("UPDATE import_scan_files SET status='重复',message=? WHERE id=?",
                                       (r.get('message', '重复'), sid)))
        except Exception as e:
            errors.append({'id': sid, 'message': str(e)})
            tx(lambda c: c.execute("UPDATE import_scan_files SET status='解析失败',message=? WHERE id=?",
                                   (str(e), sid)))
    return {'ok': True, 'imported': done, 'errors': errors}


def import_list():
    return query_all('SELECT id,file_name,relative_path,file_ext,file_size,sha256,imported_at,status,message,'
                     'quote_no,quote_date,customer_text,project_name FROM import_files '
                     'ORDER BY imported_at DESC')


# ---------- 报价单 Excel 生成 ----------
def _export_dir():
    d = os.path.join(C.DATA, 'exports')
    os.makedirs(d, exist_ok=True)
    return d


def excel_file(qid):
    from openpyxl import load_workbook
    from copy import copy
    from openpyxl.utils.cell import range_boundaries
    q = query_one('SELECT q.*,c.company,c.country,c.contact,c.email,c.address,c.phone,'
                  'p.project_name,pr.provider_name,pr.provider_info '
                  'FROM quotations q JOIN customers c ON c.id=q.customer_id '
                  'LEFT JOIN projects p ON p.id=q.project_id LEFT JOIN providers pr ON pr.id=q.provider_id '
                  'WHERE q.id=?', (qid,))
    if not q:
        return None
    items = query_all('SELECT qi.*,p.model,p.series FROM quotation_items qi '
                      'LEFT JOIN products p ON p.id=qi.product_id WHERE quotation_id=? ORDER BY item_no', (qid,))
    wb = load_workbook(C.TEMPLATE)
    ws = wb['Sheet1 (3)']
    ws._images = []          # 去掉模板内置样例图
    ws._charts = []
    extra = max(0, len(items) - 7)
    if extra:
        merges = [str(r) for r in ws.merged_cells.ranges]
        for r in merges:
            ws.unmerge_cells(r)
        ws.insert_rows(16, extra)
        for rr in range(16, 16 + extra):
            ws.row_dimensions[rr].height = ws.row_dimensions[15].height
            for cc in range(1, 9):
                src, dst = ws.cell(15, cc), ws.cell(rr, cc)
                if src.has_style:
                    dst._style = copy(src._style)
                if src.number_format:
                    dst.number_format = src.number_format
                if src.alignment:
                    dst.alignment = copy(src.alignment)
        for rg in merges:
            minc, minr, maxc, maxr = range_boundaries(rg)
            shift = extra if minr >= 16 else 0
            ws.merge_cells(start_row=minr + shift, start_column=minc,
                           end_row=maxr + shift, end_column=maxc)
    ws['A1'] = q.get('provider_info') or q.get('provider_name') or ws['A1'].value
    ws['B2'] = q['company']
    ws['H2'] = q['quote_no']
    ws['B3'] = q['contact'] or ''
    ws['B4'] = q['address'] or ''      # R4=Address（地址）
    ws['B5'] = q['phone'] or ''        # R5=Phone number（联系电话）
    ws['B7'] = q['project_name'] or ''  # R7=Project（项目名称），修正原先错写到 B4 的问题
    if re.match(r'^\d{4}-\d{2}-\d{2}', q['quote_date'] or ''):
        ws['H4'] = datetime.strptime(q['quote_date'][:10], '%Y-%m-%d')
        ws['H4'].number_format = 'yyyy/m/d'
    else:
        ws['H4'] = q['quote_date']
    detail_end = 8 + len(items)
    for r in range(9, max(detail_end, 16) + 1):
        for col in range(1, 11):
            cell = ws.cell(r, col)
            if cell.__class__.__name__ != 'MergedCell':
                cell.value = None
    for idx, x in enumerate(items, 9):
        ws.cell(idx, 1).value = idx - 8
        ws.cell(idx, 2).value = x['product_name'] or x['model'] or x['series'] or ''
        ws.cell(idx, 3).value = x['description'] or ''
        ws.cell(idx, 5).value = x['qty']
        ws.cell(idx, 6).value = x['unit']
        ws.cell(idx, 7).value = x['unit_price_usd']
        ws.cell(idx, 8).value = x['amount_usd']
        ws.cell(idx, 9).value = x['our_price_usd']
        ws.cell(idx, 10).value = x['our_amount_usd']
    total_row = 16 + extra
    ws.cell(total_row, 8).value = '=SUM(H9:H%d)' % (8 + len(items))
    path = os.path.join(_export_dir(), '%s_%s.xlsx' % (q['quote_no'], re.sub(r'[^A-Za-z0-9_-]', '_', q['company'])))
    wb.save(path)
    return path


def imported_original(qid):
    r = query_one('SELECT f.workbook_blob,f.file_name FROM import_files f JOIN quotations q ON q.quote_no=f.quote_no '
                  'WHERE q.id=? ORDER BY f.id DESC LIMIT 1', (qid,))
    if not r or not r['workbook_blob']:
        return None
    path = os.path.join(_export_dir(), re.sub(r'[^A-Za-z0-9_.-]', '_', r['file_name']))
    with open(path, 'wb') as f:
        f.write(r['workbook_blob'])
    return path
